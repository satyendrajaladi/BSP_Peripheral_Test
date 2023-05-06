# This is a sample Python script.
import io
import os
import re

import sys
import time
import serial
import openpyxl

# for sheet with timestamp
from datetime import datetime
from openpyxl.styles import colors
from openpyxl.styles import Color, PatternFill, Font, Border

print(sys.argv[0])  # prints python_script.py

# Serial Port Init Related
ser = serial.Serial('/dev/ttyUSB1', 115200)
print("Serial Instance Configuration: ", {ser})

# opening the source excel file

#filename = "/home/user/PycharmProjects/readsepthread/Automation.xlsx"
filename = sys.argv[1]
#print("xls_fileName", {filename})
# workbook instance

wb1 = openpyxl.load_workbook(filename, data_only=True)

# worksheet1 from the above workbook
ws1 = wb1.active

# calculate total number of rows and columns in source excel file

mr = ws1.max_row
mc = ws1.max_column

print('total no of rows ', {mr}, 'total no of cols ', {mc})


# utility Methods
def _readline(self):
    eol = b'\r'
    leneol = len(eol)
    line = bytearray()
    while True:
        c = ser.read(1)
        if c:
            line += c
            if line[-leneol:] == eol:
                break
        else:
            break
    return bytes(line)


# utility Methods
def copytoNewSheet():
    print("copytoNewSheet Method")
    now = datetime.now()  # current date and time

    year = now.strftime("%Y")

    month = now.strftime("%m")

    day = now.strftime("%d")

    time = now.strftime("%H %M %S")
    # print("time:", time)

    date_time_sheet = now.strftime("%d %b %Y " + time)
    target = wb1.copy_worksheet(ws1)
    target.title = str(date_time_sheet)
    # saving the destination excel file
    wb1.save(str(filename))
    wb1.close()


# serial port Methods
def close_port():
    ser.close()


# serial port Methods
def open_port():
    if ser.isOpen():
        ser.close()
        ser.open()
        time.sleep(1)
        print('SA8295P_v2.1_ft0_ADP_Air_v1.0.1_UFS_NORMAL')


def rd_frm_xl_wrt_to_ser():
    ser.flush()

    # Writing to  port
    for row in ws1.iter_rows(min_row=2, max_row=mr, min_col=2, max_col=2, values_only=True):
        for cell in row:
            # print(cell)
            ser.write(str.encode(cell))
            ser.write(str.encode("\r"))
            time.sleep(1)
    time.sleep(1)


def rd_frm_ser_wrt_xls():
    # Reading from a serial port
    expected_output_LIST = []

    mr = ws1.max_row
    index_value = 1
    pairsofdata = " "
    MAINCOLLECTOR = []
    a = None
    b = None
    while True:

        try:
            print(f"Entering Iteration - {index_value}")

            if ser.inWaiting() >= 0:

                ourStr = ser.readline().decode('utf-8').strip()
                print("ourStr: ", ourStr)

                if not pairsofdata.endswith(" ~~ "):
                    pairsofdata = pairsofdata + ourStr + " ~~ "
                    continue

                pairsofdata = pairsofdata + ourStr

                print("This is the pair of data")
                print(pairsofdata)

                count = pairsofdata.count(" ~~ ")
                print("Count of Tildes:", count)

                if count == 1:
                    MAINLIST = pairsofdata.split(" ~~ ")
                    print("MAINLIST :", MAINLIST)
                    print()
                if len(MAINLIST) == 2:

                    ####### Extracting the second element from MAINLIST

                    expectedOutput = ws1.cell(row=index_value + 1, column=3).value
                    print("This is expected value:")
                    print(expectedOutput)
                    if MAINLIST[1] == expectedOutput:

                        print(f"I have found the data {MAINLIST[1]}")
                        Res = ws1.cell(row=index_value + 1, column=4)

                        Res.value = 'Pass'

                        my_green = openpyxl.styles.colors.Color(rgb='00FF00')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)
                        Res.fill = my_fill

                        Reason = ws1.cell(row=index_value + 1, column=5)
                        Reason.value = 'Both the Actual output and Expected Output Matches,Hence TC is Pass.'

                    else:
                        Res = ws1.cell(row=index_value + 1, column=4)
                        Res.value = 'Fail'
                        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
                        Res.fill = my_fill
                        Reason = ws1.cell(row=index_value + 1, column=5)
                        reason_str = ('ExpectedOutput is: ', str(expectedOutput), 'Actual Output is: ', str(ourStr))
                        Reason.value = str(reason_str)

                    ####### Post Processing and clearing data
                    pairsofdata = " "
                    index_value += 1

                    if index_value == mr:
                        break
                    continue

                continue

            if index_value == mr + 15:
                break

            index_value += 1

        except Exception as e:
            print(" Interrupt Error is here  :-- ")
            print(e)
            break

    wb1.save(filename)
    wb1.close()


def rd_ln_by_ln():
    index_value = 1
    remoteIndexing = 0
    pairsofdata = " "
    MAINCOLLECTOR = []
    a = None
    b = None

    while True:
        print("index_value: ", index_value)
        CURRENT_COMMAND = ws1.cell(row=index_value + 1, column=2).value
        print("CURRENT_COMMAND:", CURRENT_COMMAND)
        NEXT_COMMAND = ws1.cell(row=index_value + 2, column=2).value
        print("NEXT_COMMAND:", NEXT_COMMAND)

        if ser.inWaiting() >= 0:
            ourStr = ser.readline().decode('utf-8').strip()
            print("ourStr: ", ourStr)

            if ourStr == CURRENT_COMMAND:
                print("SAME")
                index_value += 1
                continue

            if ourStr != NEXT_COMMAND:
                MAINCOLLECTOR.append(ourStr + "\n")
                remoteIndexing += 1

            index_value += 1
            ROW_NUMBER = index_value + 1 - remoteIndexing


if __name__ == '__main__':
    open_port()
    rd_frm_xl_wrt_to_ser()
    rd_frm_ser_wrt_xls()
    copytoNewSheet()
    close_port()
